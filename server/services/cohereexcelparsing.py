#!/usr/bin/env python3

import argparse
import re
import os
import sys
import json
import logging
import time
from datetime import datetime
from typing import List, Tuple, Dict, Optional, Any, Union
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import cohere
import uuid
import traceback

import pandas as pd
from scipy.spatial.distance import cosine

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('price_matching.log', mode='a')
    ]
)
logger = logging.getLogger(__name__)

# Enhanced preprocessing with better synonym mapping and stop words
SYNONYM_MAP = {
    # Building materials
    "bricks": "brick", "brickwork": "brick", "blocks": "brick", "blockwork": "brick",
    "masonry": "brick", "stonework": "stone", "tiles": "tile", "tiling": "tile",
    
    # Concrete and cement
    "cement": "concrete", "concrete": "concrete", "mortar": "concrete",
    "grout": "concrete", "screed": "concrete", "render": "concrete",
    
    # Foundation work
    "footing": "foundation", "footings": "foundation", "foundations": "foundation",
    "basement": "foundation", "substructure": "foundation",
    
    # Excavation and earthwork
    "excavation": "excavate", "excavations": "excavate", "excavate": "excavate", 
    "dig": "excavate", "digging": "excavate", "earthwork": "excavate",
    "trenching": "excavate", "grading": "excavate",
    
    # Installation and construction
    "installation": "install", "installing": "install", "installed": "install",
    "construction": "build", "building": "build", "erection": "install",
    "assembly": "install", "fitting": "install",
    
    # Demolition and removal
    "demolition": "demolish", "demolish": "demolish", "demolishing": "demolish", 
    "remove": "demolish", "removal": "demolish", "strip": "demolish",
    "break": "demolish", "dismantle": "demolish",
    
    # Supply and provision
    "supply": "provide", "supplies": "provide", "providing": "provide",
    "furnish": "provide", "deliver": "provide", "procurement": "provide",
    
    # Finishes
    "painting": "paint", "plastering": "plaster", "flooring": "floor",
    "roofing": "roof", "cladding": "clad", "insulation": "insulate",
    
    # MEP (Mechanical, Electrical, Plumbing)
    "electrical": "electric", "plumbing": "plumb", "hvac": "ventilation",
    "heating": "heat", "cooling": "cool", "ventilation": "ventilate",
    
    # Structural
    "reinforcement": "reinforce", "steelwork": "steel", "formwork": "form",
    "shuttering": "shutter", "framework": "frame"
}

STOP_WORDS = {
    "the", "and", "of", "to", "in", "for", "on", "at", "by", "from", "with",
    "a", "an", "be", "is", "are", "as", "it", "its", "into", "or", "this",
    "that", "will", "shall", "would", "could", "should", "may", "might",
    "per", "each", "all", "any", "some", "no", "not", "only", "such",
    "than", "too", "very", "can", "had", "her", "was", "one", "our", "out",
    "day", "get", "has", "him", "his", "how", "man", "new", "now", "old",
    "see", "two", "way", "who", "boy", "did", "use", "she", "they", "we"
}

# Enhanced configuration
EMBED_BATCH = 90  # Slightly reduced for stability
EMBED_MODEL = "embed-v4.0"
OUTPUT_DIMENSION = 1536
SIMILARITY_THRESHOLD = 0.3  # Minimum similarity score to consider a match
MAX_RETRIES = 3
RETRY_DELAY = 1.0

class ProgressTracker:
    """Track and report progress during processing"""
    
    def __init__(self, total_steps: int = 100):
        self.total_steps = total_steps
        self.current_step = 0
        self.start_time = datetime.now()
    
    def update(self, step: int, message: str = ""):
        self.current_step = step
        percentage = (step / self.total_steps) * 100
        elapsed = datetime.now() - self.start_time
        
        progress_msg = f"Progress: {percentage:.1f}% ({step}/{self.total_steps})"
        if message:
            progress_msg += f" - {message}"
        
        logger.info(progress_msg)
        print(f"PROGRESS: {percentage:.1f}%", flush=True)
    
    def complete(self, message: str = "Processing completed"):
        elapsed = datetime.now() - self.start_time
        logger.info(f"{message} in {elapsed.total_seconds():.2f} seconds")

def enhanced_preprocess(text: str, synonym_map: Dict[str, str], stop_words: set) -> str:
    """Enhanced text preprocessing with better normalization"""
    if not text or text is None or (hasattr(text, '__len__') and len(str(text).strip()) == 0):
        return ""
    
    # Convert to string and lowercase
    s = str(text).lower().strip()
    
    # Remove special characters but preserve spaces and alphanumeric
    s = re.sub(r"[^\w\s]", " ", s)
    
    # Normalize units and measurements
    s = re.sub(r"\b\d+(?:\.\d+)?\s*(mm|cm|m|inch|in|ft|feet|yard|yd)\b", " UNIT ", s)
    
    # Normalize numbers
    s = re.sub(r"\b\d+(?:\.\d+)?\b", " NUM ", s)
    
    # Remove extra whitespace
    s = re.sub(r"\s+", " ", s).strip()
    
    # Apply synonyms and stemming
    words = []
    for word in s.split():
        # Apply synonym mapping
        word = synonym_map.get(word, word)
        
        # Simple stemming for common suffixes
        if len(word) > 4:
            word = re.sub(r"(ings|ing|ed|es|s|tion|sion)$", "", word)
        
        # Filter out stop words and very short words
        if word not in stop_words and len(word) > 2:
            words.append(word)
    
    return " ".join(words)

def embed_texts_with_retry(client: cohere.Client, texts: List[str], input_type: str = "search_document") -> np.ndarray:
    """Embed texts with retry logic and better error handling"""
    embeddings = []
    
    for i in range(0, len(texts), EMBED_BATCH):
        batch = texts[i:i+EMBED_BATCH]
        batch_num = (i // EMBED_BATCH) + 1
        total_batches = (len(texts) + EMBED_BATCH - 1) // EMBED_BATCH
        
        logger.info(f"Processing embedding batch {batch_num}/{total_batches} ({len(batch)} items)")
        
        # Add delay between batches to respect rate limits (especially for trial accounts)
        if batch_num > 1:
            rate_limit_delay = 2.0  # 2 second delay between batches
            logger.debug(f"Waiting {rate_limit_delay}s to respect rate limits...")
            time.sleep(rate_limit_delay)
        
        for attempt in range(MAX_RETRIES):
            try:
                resp = client.embed(
                    texts=batch,
                    model=EMBED_MODEL,
                    input_type=input_type,
                    embedding_types=["float"]
                )
                
                # Debug: Log response structure
                logger.debug(f"API response type: {type(resp)}")
                logger.debug(f"API response has embeddings: {hasattr(resp, 'embeddings')}")
                
                # Extract embeddings with proper validation for Cohere v4.0
                if hasattr(resp, 'embeddings') and resp.embeddings:
                    # For Cohere v4.0, embeddings is an EmbedByTypeResponseEmbeddings object
                    batch_embeddings_raw = resp.embeddings
                    logger.debug(f"Batch embeddings type: {type(batch_embeddings_raw)}")
                    
                    # Try to access the float embeddings
                    embeddings_list = []
                    
                    # Method 1: Try to access 'float' attribute (Cohere v4.0)
                    float_embeddings = getattr(batch_embeddings_raw, 'float', None)
                    if float_embeddings is not None:
                        logger.debug(f"Found float embeddings (v4.0 format)")
                        logger.debug(f"Float embeddings type: {type(float_embeddings)}")
                        logger.debug(f"Float embeddings length: {len(float_embeddings)}")
                        
                        # Each embedding should be a list of floats
                        for emb in float_embeddings:
                            if isinstance(emb, list) and len(emb) > 100:  # Validate dimension
                                embeddings_list.append(emb)
                            else:
                                logger.warning(f"Invalid embedding format: {type(emb)}, length: {len(emb) if hasattr(emb, '__len__') else 'No length'}")
                    
                    # Method 2: Try direct iteration (older formats or direct list)
                    elif isinstance(batch_embeddings_raw, list):
                        logger.debug(f"Found direct list of embeddings (older format)")
                        for emb in batch_embeddings_raw:
                            if isinstance(emb, list) and len(emb) > 100:
                                embeddings_list.append(emb)
                            else:
                                logger.warning(f"Invalid direct embedding: {type(emb)}, length: {len(emb) if hasattr(emb, '__len__') else 'No length'}")
                    
                    # Method 3: Try iterating and checking attributes
                    elif hasattr(batch_embeddings_raw, '__iter__'):
                        logger.debug(f"Trying attribute-based extraction")
                        for emb in batch_embeddings_raw:
                            if isinstance(emb, list) and len(emb) > 100:
                                embeddings_list.append(emb)
                            else:
                                # Try different attribute names
                                emb_float = getattr(emb, 'float', None)
                                values = getattr(emb, 'values', None)
                                embedding = getattr(emb, 'embedding', None)
                                
                                if emb_float is not None and isinstance(emb_float, list) and len(emb_float) > 100:
                                    embeddings_list.append(emb_float)
                                elif values is not None and hasattr(values, '__iter__') and len(list(values)) > 100:
                                    embeddings_list.append(list(values))
                                elif embedding is not None and hasattr(embedding, '__iter__') and len(list(embedding)) > 100:
                                    embeddings_list.append(list(embedding))
                                else:
                                    logger.warning(f"Could not extract valid embedding from: {type(emb)}")
                                    continue
                    
                    if not embeddings_list:
                        logger.error(f"No valid embeddings extracted from batch")
                        logger.error(f"Raw embeddings type: {type(batch_embeddings_raw)}")
                        logger.error(f"Raw embeddings type: {type(batch_embeddings_raw)}")
                        # Safe length check for type checker
                        try:
                            # Type-safe length check
                            raw_len = getattr(batch_embeddings_raw, '__len__', None)
                            if raw_len is not None:
                                logger.error(f"Raw embeddings length: {raw_len()}")
                        except:
                            pass
                        raise ValueError(f"No valid embeddings extracted from batch. Raw type: {type(batch_embeddings_raw)}")
                    
                    logger.debug(f"Extracted {len(embeddings_list)} embeddings from batch")
                    if embeddings_list:
                        logger.debug(f"First embedding length: {len(embeddings_list[0])}")
                        logger.debug(f"Expected embedding dimension: {OUTPUT_DIMENSION}")
                    
                    # Validate all embeddings have same dimension
                    first_dim = len(embeddings_list[0])
                    for idx, emb in enumerate(embeddings_list):
                        if len(emb) != first_dim:
                            logger.error(f"Inconsistent embedding dimensions: embedding {idx} has {len(emb)}, expected {first_dim}")
                            raise ValueError(f"Inconsistent embedding dimensions in batch")
                    
                    embeddings.extend(embeddings_list)
                    
                else:
                    raise ValueError("No embeddings found in API response")
                
                break
                
            except Exception as e:
                logger.warning(f"Embedding attempt {attempt + 1} failed: {str(e)}")
                if attempt == MAX_RETRIES - 1:
                    raise Exception(f"Failed to get embeddings after {MAX_RETRIES} attempts: {str(e)}")
                
                time.sleep(RETRY_DELAY * (attempt + 1))
    
    # Convert to numpy array and validate final shape
    try:
        logger.info(f"Total embeddings collected: {len(embeddings)}")
        if embeddings:
            logger.info(f"First embedding sample length: {len(embeddings[0])}")
            logger.info(f"Last embedding sample length: {len(embeddings[-1])}")
            
            # Check for consistency in embedding dimensions
            embedding_lengths = [len(emb) for emb in embeddings[:5]]  # Check first 5
            logger.info(f"First 5 embedding lengths: {embedding_lengths}")
            
            # Validate all embeddings have consistent dimensions
            if len(set(embedding_lengths)) > 1:
                raise ValueError(f"Inconsistent embedding dimensions: {embedding_lengths}")
        
        embeddings_array = np.array(embeddings, dtype=np.float32)
        logger.info(f"Created embeddings array with shape: {embeddings_array.shape}")
        
        # Ensure we have a 2D array with proper dimensions
        if len(embeddings_array.shape) != 2:
            raise ValueError(f"Expected 2D embeddings array, got shape {embeddings_array.shape}")
        
        # Validate embedding dimension (should be 1024+ for embed-v4.0)
        if embeddings_array.shape[1] < 100:  # Reasonable minimum dimension check
            raise ValueError(f"Embedding dimension too small: {embeddings_array.shape[1]}, expected ~1024")
        
        return embeddings_array
        
    except Exception as e:
        logger.error(f"Failed to create embeddings array: {str(e)}")
        logger.error(f"Embeddings list length: {len(embeddings)}")
        if embeddings:
            logger.error(f"First embedding type: {type(embeddings[0])}")
            logger.error(f"First embedding length: {len(embeddings[0]) if hasattr(embeddings[0], '__len__') else 'No length'}")
            if hasattr(embeddings[0], '__len__') and len(embeddings[0]) > 5:
                logger.error(f"First embedding sample: {embeddings[0][:5]}")
            else:
                logger.error(f"First embedding: {embeddings[0]}")
        raise ValueError(f"Failed to create valid embeddings array: {str(e)}")

def load_pricelist_enhanced(path: str) -> Tuple[List[str], List[float], List[str], List[str]]:
    """Load pricelist with enhanced validation and metadata"""
    logger.info(f"Loading pricelist from: {path}")
    
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        
        if ws is None:
            raise ValueError("No active worksheet found in pricelist file")
        
        descriptions = []
        rates = []
        units = []
        
        # Find header row and column positions
        header_row = 1
        id_col = 0
        desc_col = 1
        rate_col = 2
        unit_col = 3
        
        for row in range(1, min(6, ws.max_row + 1)):
            for col_idx, cell in enumerate(ws[row]):
                if cell.value:
                    cell_val = str(cell.value).lower()
                    if 'id' in cell_val:
                        id_col = col_idx
                        header_row = row
                    elif 'description' in cell_val:
                        desc_col = col_idx
                        header_row = row
                    elif 'rate' in cell_val:
                        rate_col = col_idx
                    elif 'unit' in cell_val:
                        unit_col = col_idx
        
        logger.info(f"Found header row at: {header_row}, columns - ID: {id_col}, Desc: {desc_col}, Rate: {rate_col}, Unit: {unit_col}")
        
        # Process data rows
        valid_items = 0
        ids = []
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if len(row) < max(desc_col + 1, rate_col + 1):
                continue
                
            item_id = row[id_col] if len(row) > id_col else None
            desc = row[desc_col] if len(row) > desc_col else None
            rate = row[rate_col] if len(row) > rate_col else None
            unit = row[unit_col] if len(row) > unit_col else "each"
            
            # Validate data
            if desc and rate is not None:
                try:
                    # Handle different cell value types
                    if isinstance(rate, (int, float)):
                        rate_float = float(rate)
                    elif isinstance(rate, str):
                        rate_float = float(rate)
                    else:
                        # Skip non-numeric values
                        logger.warning(f"Non-numeric rate at row {row_num}: {rate}")
                        continue
                        
                    if rate_float > 0:  # Only positive rates
                        descriptions.append(str(desc).strip())
                        rates.append(rate_float)
                        units.append(str(unit) if unit else "each")
                        ids.append(str(item_id) if item_id else "")
                        valid_items += 1
                except (ValueError, TypeError):
                    logger.warning(f"Invalid rate at row {row_num}: {rate}")
                    continue
        
        wb.close()
        logger.info(f"Loaded {valid_items} valid price items from pricelist")
        
        if valid_items == 0:
            raise ValueError("No valid price items found in pricelist")
        
        return descriptions, rates, units, ids
        
    except Exception as e:
        logger.error(f"Error loading pricelist: {str(e)}")
        raise

def find_headers_enhanced(ws) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Find headers in worksheet with enhanced detection and validation.
    Now includes content validation to ensure selected columns contain meaningful data.
    
    Returns: (header_row, description_column, quantity_column)
    """
    header_row = None
    desc_col = None
    qty_col = None
    
    logger.info(f"Enhanced header detection starting (max rows: {ws.max_row}, max cols: {ws.max_column})")
    
    # Track potential description columns with quality scores
    description_candidates = []
    
    # Look for header patterns in the first 15 rows
    for row_num in range(1, min(16, ws.max_row + 1)):
        row = ws[row_num]
        
        header_indicators = 0
        potential_desc_col = None
        potential_qty_col = None
        
        for col_num, cell in enumerate(row, start=1):
            if not cell.value:
                continue
                
            cell_value = str(cell.value).lower().strip()
            
            # DESCRIPTION COLUMN - Very comprehensive patterns
            desc_patterns = [
                'description', 'desc', 'item', 'work', 'activity', 'task', 'operation',
                'specification', 'spec', 'details', 'particulars', 'trade', 'element',
                'component', 'material', 'labor', 'labour', 'service', 'works',
                'ref', 'reference', 'code', 'product', 'title', 'name', 'category'
            ]
            
            if any(pattern in cell_value for pattern in desc_patterns):
                # Avoid false positives
                if not any(exclude in cell_value for exclude in ['note', 'remark', 'comment', 'total', 'sum', 'page']):
                    # Validate this column actually contains meaningful descriptions
                    quality = validate_description_column(ws, row_num, col_num)
                    description_candidates.append({
                        'row_num': row_num,
                        'col_num': col_num,
                        'quality': quality,
                        'header_name': str(cell.value).strip()
                    })
                    logger.info(f"Found potential description column at row {row_num}, col {col_num}: '{cell.value}' (quality: {quality})")
            
            # QUANTITY COLUMN - Ultra comprehensive patterns
            qty_patterns = [
                # Standard patterns
                'qty', 'quantity', 'quan', 'qnty', 'qtty', 'quantities',
                'amount', 'amt', 'number', 'no.', 'no', 'nr', 'num', '#',
                'units', 'unit', 'count', 'cnt', 'total', 'sum',
                # Measurement patterns
                'volume', 'area', 'length', 'width', 'height', 'depth',
                'm2', 'm²', 'm3', 'm³', 'sqm', 'cbm', 'linear', 'lm',
                'sq.m', 'cu.m', 'square', 'cubic', 'metres', 'meters',
                # Unit patterns
                'each', 'ea', 'pcs', 'pieces', 'items', 'nos', 'numbers',
                'kg', 'tonnes', 'tons', 'litres', 'liters', 'hours', 'hrs',
                # Other patterns
                'extent', 'measure', 'size', 'dimension', 'scope'
            ]
            
            if any(pattern in cell_value for pattern in qty_patterns):
                # Avoid false positives
                if not any(exclude in cell_value for exclude in ['description', 'note', 'remark', 'comment', 'page']):
                    potential_qty_col = col_num
                    header_indicators += 1
                    logger.info(f"Found QUANTITY pattern in col {col_num}: '{cell.value}'")
            
            # Other header indicators
            other_patterns = ['rate', 'price', 'cost', 'value', 'total', 'unit', 'measure']
            if any(pattern in cell_value for pattern in other_patterns):
                header_indicators += 0.5
        
        # Check if we found a good quantity column for this row
        if potential_qty_col and header_indicators >= 1:
            qty_col = potential_qty_col
            # Keep the row number for potential use
            if not header_row:
                header_row = row_num
    
    # Select the best description column based on quality
    if description_candidates:
        best_candidate = max(description_candidates, key=lambda x: x['quality'])
        if best_candidate['quality'] >= 3:  # Minimum quality threshold
            desc_col = best_candidate['col_num']
            header_row = best_candidate['row_num']
            logger.info(f"✅ Selected description column {desc_col}: '{best_candidate['header_name']}' (quality: {best_candidate['quality']})")
        else:
            logger.warning(f"Best description candidate has low quality ({best_candidate['quality']}), trying fallback...")
    
    # If no quantity column found by name, search more aggressively
    if desc_col and not qty_col and header_row:
        logger.warning("No quantity column found by pattern matching, performing ADAPTIVE SEARCH...")
        
        # Expanded search range - check more columns
        search_range = list(range(1, min(ws.max_column + 1, 20)))  # Search first 20 columns
        logger.info(f"Searching for numeric columns in range: {search_range}")
        
        best_qty_col = None
        best_score = 0
        
        for test_col in search_range:
            if test_col == desc_col:
                continue
                
            # Use the new validation function
            quality = validate_quantity_column(ws, header_row, test_col)
            if quality > best_score and quality >= 2:  # Minimum quality threshold
                best_score = quality
                best_qty_col = test_col
                header_name = ws.cell(row=header_row, column=test_col).value or f"Column_{test_col}"
                logger.info(f"Candidate quantity column {test_col} ('{header_name}'): quality={quality}")
        
        if best_qty_col:
            qty_col = best_qty_col
            header_name = ws.cell(row=header_row, column=qty_col).value
            logger.info(f"AUTO-DETECTED quantity column {qty_col} ('{header_name}') with quality {best_score}")
    
    # Final validation and fallback
    if not desc_col:
        logger.warning("No description column found! Trying fallback detection...")
        # Try to find the first text-heavy column
        for col_num in range(1, min(10, ws.max_column + 1)):
            quality = validate_description_column(ws, 1, col_num)
            if quality >= 2:  # Lower threshold for fallback
                desc_col = col_num
                header_row = 1
                logger.info(f"FALLBACK: Using column {col_num} as description column (quality: {quality})")
                break
    
    logger.info(f"=== FINAL DETECTION RESULT ===")
    logger.info(f"Header Row: {header_row}")
    logger.info(f"Description Column: {desc_col}")
    logger.info(f"Quantity Column: {qty_col}")
    logger.info(f"=======================")
    
    return header_row, desc_col, qty_col

def validate_description_column(ws, header_row: int, col_num: int) -> int:
    """
    Validate that a column actually contains meaningful descriptions.
    Returns a quality score (higher is better, 0-10 scale).
    """
    quality = 0
    text_rows = 0
    meaningful_rows = 0
    total_length = 0
    numeric_rows = 0
    short_rows = 0
    
    # Check up to 20 data rows after the header
    max_rows = min(header_row + 21, ws.max_row + 1)
    for row_num in range(header_row + 1, max_rows):
        cell = ws.cell(row=row_num, column=col_num)
        if not cell.value:
            continue
        
        cell_value = str(cell.value).strip()
        if len(cell_value) == 0:
            continue
        
        text_rows += 1
        total_length += len(cell_value)
        
        # Check if it's mostly numeric (bad for descriptions)
        import re
        if re.match(r'^\d+(\.\d+)?$', cell_value):
            numeric_rows += 1
        
        # Check if it's too short (bad for descriptions)
        if len(cell_value) <= 3:
            short_rows += 1
        
        # Check for meaningful descriptive content
        if (len(cell_value) > 10 and
            re.search(r'[a-zA-Z]', cell_value) and
            not re.match(r'^\d+$', cell_value)):
            meaningful_rows += 1
    
    if text_rows == 0:
        return 0
    
    # Calculate quality score
    average_length = total_length / text_rows
    meaningful_ratio = meaningful_rows / text_rows
    numeric_ratio = numeric_rows / text_rows
    short_ratio = short_rows / text_rows
    
    # Start with base score
    quality = 1
    
    # Bonus for meaningful content
    if meaningful_ratio > 0.7:
        quality += 5
    elif meaningful_ratio > 0.5:
        quality += 3
    elif meaningful_ratio > 0.3:
        quality += 1
    
    # Bonus for good average length
    if average_length > 30:
        quality += 3
    elif average_length > 15:
        quality += 2
    elif average_length > 5:
        quality += 1
    
    # Penalty for too many numeric values
    if numeric_ratio > 0.8:
        quality -= 5
    elif numeric_ratio > 0.5:
        quality -= 3
    elif numeric_ratio > 0.3:
        quality -= 1
    
    # Penalty for too many short values
    if short_ratio > 0.8:
        quality -= 3
    elif short_ratio > 0.5:
        quality -= 2
    
    logger.debug(f"Column {col_num} validation: {text_rows} rows, avg length: {average_length:.1f}, meaningful: {meaningful_ratio:.2f}, numeric: {numeric_ratio:.2f}, quality: {quality}")
    
    return max(0, quality)

def validate_quantity_column(ws, header_row: int, col_num: int) -> int:
    """
    Validate that a column contains numeric quantities.
    Returns a quality score (higher is better, 0-10 scale).
    """
    quality = 0
    numeric_rows = 0
    total_rows = 0
    positive_rows = 0
    decimal_rows = 0
    
    # Check up to 20 data rows after the header
    max_rows = min(header_row + 21, ws.max_row + 1)
    for row_num in range(header_row + 1, max_rows):
        cell = ws.cell(row=row_num, column=col_num)
        if not cell.value:
            continue
        
        cell_value = str(cell.value).strip()
        if len(cell_value) == 0:
            continue
        
        total_rows += 1
        
        # Check if it's numeric
        try:
            value = float(cell_value.replace(',', ''))
            numeric_rows += 1
            
            if value > 0:
                positive_rows += 1
            
            if '.' in cell_value or ',' in cell_value:
                decimal_rows += 1
        except (ValueError, TypeError):
            pass
    
    if total_rows == 0:
        return 0
    
    numeric_ratio = numeric_rows / total_rows
    positive_ratio = positive_rows / total_rows
    
    # Start with base score if mostly numeric
    if numeric_ratio > 0.7:
        quality += 3
    elif numeric_ratio > 0.5:
        quality += 2
    elif numeric_ratio > 0.3:
        quality += 1
    
    # Bonus for positive values (quantities should be positive)
    if positive_ratio > 0.5:
        quality += 2
    elif positive_ratio > 0.3:
        quality += 1
    
    # Bonus for having some decimal values (common in quantities)
    if decimal_rows > 0:
        quality += 1
    
    return max(0, quality)

def add_result_columns_with_formatting(ws, header_row: int, max_col: int):
    """Add result columns with proper formatting"""
    # Define new column headers
    new_headers = ['Matched Description', 'Matched Rate', 'Total Amount', 'Similarity Score', 'Match Quality', 'Matched ID', 'Unit']
    
    # Add headers with formatting
    for i, header in enumerate(new_headers, start=1):
        col_num = max_col + i
        cell = ws.cell(row=header_row, column=col_num, value=header)
        
        # Apply header formatting
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    return len(new_headers)

def calculate_match_quality(similarity: float) -> str:
    """Determine match quality based on similarity score"""
    if similarity >= 0.9:
        return "Excellent"
    elif similarity >= 0.8:
        return "Very Good"
    elif similarity >= 0.7:
        return "Good"
    elif similarity >= 0.6:
        return "Fair"
    elif similarity >= 0.5:
        return "Poor"
    else:
        return "Very Poor"

def hierarchical_match_scoring(inquiry_item: dict, price_descriptions: List[str], price_rates: List[float], 
                             price_units: List[str], similarities: np.ndarray) -> Tuple[int, float, dict]:
    """
    Implement hierarchical matching logic:
    1. Category identification
    2. Item description matching
    3. Related keywords
    4. Related phrases
    5. Unit matching
    6. Full context matching
    
    Returns: (best_index, best_score, match_details)
    """
    inquiry_desc = inquiry_item['description'].lower()
    inquiry_enhanced = inquiry_item['enhanced_description'].lower()
    
    # Extract potential unit from inquiry
    import re
    unit_patterns = [r'\b(m2|m²|sqm|square\s*meter?s?)\b', r'\b(m3|m³|cubic\s*meter?s?)\b', 
                    r'\b(kg|kilogram?s?)\b', r'\b(ton?s?|tonne?s?)\b', r'\b(liter?s?|litre?s?|l)\b',
                    r'\b(piece?s?|pcs?|each|no\.?|number?s?)\b', r'\b(hour?s?|hr?s?)\b',
                    r'\b(day?s?)\b', r'\b(week?s?)\b', r'\b(month?s?)\b']
    
    inquiry_units = []
    for pattern in unit_patterns:
        matches = re.findall(pattern, inquiry_enhanced)
        inquiry_units.extend(matches)
    
    # Score each price item using hierarchical criteria
    enhanced_scores = []
    
    for i, base_similarity in enumerate(similarities):
        score_components = {
            'base_similarity': base_similarity,
            'category_boost': 0.0,
            'keyword_boost': 0.0,
            'phrase_boost': 0.0,
            'unit_boost': 0.0,
            'context_boost': 0.0
        }
        
        price_desc = price_descriptions[i].lower()
        price_unit = price_units[i].lower() if i < len(price_units) else ''
        
        # 1. Category identification boost
        if inquiry_item.get('head_title'):
            head_title = inquiry_item['head_title'].lower()
            if any(word in price_desc for word in head_title.split() if len(word) > 3):
                score_components['category_boost'] = 0.1
        
        # 2. Item description keyword matching
        inquiry_words = set(word for word in inquiry_desc.split() if len(word) > 2)
        price_words = set(word for word in price_desc.split() if len(word) > 2)
        common_words = inquiry_words.intersection(price_words)
        if common_words:
            keyword_ratio = len(common_words) / len(inquiry_words) if inquiry_words else 0
            score_components['keyword_boost'] = keyword_ratio * 0.15
        
        # 3. Related phrases boost
        inquiry_phrases = [inquiry_desc[j:j+10] for j in range(len(inquiry_desc)-9)]
        for phrase in inquiry_phrases:
            if len(phrase) > 5 and phrase in price_desc:
                score_components['phrase_boost'] = min(score_components['phrase_boost'] + 0.05, 0.1)
        
        # 4. Unit matching boost
        if inquiry_units and price_unit:
            for inq_unit in inquiry_units:
                if inq_unit in price_unit or price_unit in inq_unit:
                    score_components['unit_boost'] = 0.1
                    break
        
        # Calculate final enhanced score
        total_boost = sum(score_components[k] for k in score_components if k != 'base_similarity')
        enhanced_score = base_similarity + total_boost
        enhanced_score = min(enhanced_score, 1.0)  # Cap at 1.0
        
        enhanced_scores.append((enhanced_score, score_components))
    
    # Find best match
    best_idx = int(np.argmax([score[0] for score in enhanced_scores]))  # Convert numpy int to Python int
    best_score = enhanced_scores[best_idx][0]
    match_details = enhanced_scores[best_idx][1]
    
    return best_idx, best_score, match_details

def process_all_sheets(workbook_path: str, pricelist_df: pd.DataFrame, job_id: str, client: cohere.Client) -> Optional[str]:
    """Process all sheets in the workbook with adaptive detection"""
    try:
        workbook = load_workbook(workbook_path, data_only=True)
        all_items = []
        total_processed = 0
        
        logger.info(f"=== PROCESSING WORKBOOK WITH {len(workbook.sheetnames)} SHEETS ===")
        
        for sheet_name in workbook.sheetnames:
            logger.info(f"\n=== PROCESSING SHEET: {sheet_name} ===")
            try:
                ws = workbook[sheet_name]
                
                # Skip empty or very small sheets
                if ws.max_row < 3 or ws.max_column < 2:
                    logger.info(f"Skipping sheet '{sheet_name}' - too small ({ws.max_row} rows, {ws.max_column} cols)")
                    continue
                
                # Use enhanced header detection
                header_row, desc_col, qty_col = find_headers_enhanced(ws)
                
                if not desc_col:
                    logger.warning(f"No description column found in sheet '{sheet_name}', trying basic fallback...")
                    # Last resort: use first column as description
                    header_row = 1
                    desc_col = 1
                    logger.info(f"Using fallback: column 1 as description in sheet '{sheet_name}'")
                
                # Process items from this sheet
                if header_row is not None and desc_col is not None:
                    sheet_items = extract_items_from_sheet(ws, header_row, desc_col, qty_col, sheet_name)
                else:
                    logger.warning(f"Invalid header detection in sheet '{sheet_name}' - skipping")
                    continue
                
                if sheet_items:
                    all_items.extend(sheet_items)
                    total_processed += len(sheet_items)
                    logger.info(f"Sheet '{sheet_name}' contributed {len(sheet_items)} items")
                else:
                    logger.warning(f"No items extracted from sheet '{sheet_name}'")
                
            except Exception as e:
                logger.error(f"Error processing sheet '{sheet_name}': {e}")
                continue
        
        logger.info(f"\n=== WORKBOOK PROCESSING COMPLETE ===")
        logger.info(f"Total items extracted from all sheets: {total_processed}")
        
        if not all_items:
            logger.error("No items found in any sheet of the workbook!")
            return None
        
        # Create DataFrame from all items
        items_df = pd.DataFrame(all_items)
        logger.info(f"Created DataFrame with {len(items_df)} total items")
        
        # Process matches
        return process_item_matching(items_df, pricelist_df, job_id, client)
        
    except Exception as e:
        logger.error(f"Error processing workbook: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def extract_items_from_sheet(ws, header_row: int, desc_col: int, qty_col: Optional[int], sheet_name: str) -> List[Dict]:
    """Extract items from a single sheet with robust BOQ format handling and deduplication"""
    import re
    items = []
    section_stack = []  # Track hierarchy of sections
    seen_descriptions = set()  # Track unique descriptions to avoid duplicates
    
    logger.info(f"Extracting items from sheet '{sheet_name}' starting at row {header_row + 1}")
    logger.info(f"Description column: {desc_col}, Quantity column: {qty_col}")
    
    # Check if there are any data rows
    data_rows_found = 0
    total_rows_scanned = 0
    unique_items_added = 0
    
    for row_num in range(header_row + 1, ws.max_row + 1):
        total_rows_scanned += 1
        desc_cell = ws.cell(row=row_num, column=desc_col)
        
        if not desc_cell.value or str(desc_cell.value).strip() == "":
            continue
            
        description = str(desc_cell.value).strip()
        
        # Skip completely empty rows
        if len(description) < 1:
            continue
        
        # Get quantity - more flexible handling with scanning
        quantity = None
        if qty_col:
            qty_cell = ws.cell(row=row_num, column=qty_col)
            if qty_cell.value is not None:
                qty_str = str(qty_cell.value).strip()
                try:
                    # Handle various formats: "5.00", "5,000.00", "5.00 m2", etc.
                    number_match = re.search(r'[\d,]+\.?\d*', qty_str.replace(',', ''))
                    if number_match:
                        quantity = float(number_match.group().replace(',', ''))
                        if quantity > 0:
                            data_rows_found += 1
                except:
                    pass
        
        # If no quantity column detected, scan nearby columns for numeric values
        if quantity is None and qty_col is None:
            for scan_col in range(max(1, desc_col - 2), min(ws.max_column + 1, desc_col + 5)):
                if scan_col == desc_col:
                    continue
                try:
                    scan_cell = ws.cell(row=row_num, column=scan_col)
                    if scan_cell.value is not None:
                        test_qty = float(scan_cell.value)
                        if test_qty > 0:
                            quantity = test_qty
                            data_rows_found += 1
                except (ValueError, TypeError):
                    continue
        
        # Enhanced section detection for BOQ formats
        is_section = detect_boq_section_header(description, quantity, row_num, ws)
        
        if is_section:
            update_section_stack(section_stack, description)
            logger.debug(f"BOQ Section detected: {description} -> Stack: {section_stack}")
            continue
        
        # Minimal filtering - only skip obviously problematic items
        # Accept all items that have any description to maximize extraction
        if len(description.strip()) < 1:
            continue
        
        # Only skip items that are clearly not items at all
        if description.lower().strip() in ['', 'total', 'subtotal', 'grand total', 'page']:
            logger.debug(f"Skipping row {row_num}: obvious non-item: '{description}'")
            continue
        
        # Default quantity for rate-based items
        if quantity is None or quantity <= 0:
            quantity = 1.0
            logger.debug(f"Row {row_num}: Defaulted quantity to 1.0 for: '{description[:30]}...'")
        
        # Build enhanced description with BOQ section context
        enhanced_description = build_enhanced_description(description, section_stack)
        
        # Create normalized key for deduplication
        normalized_desc = normalize_description_for_dedup(description)
        
        # Skip if we've seen this description before (deduplication)
        if normalized_desc in seen_descriptions:
            logger.debug(f"Skipping duplicate row {row_num}: '{description[:30]}...'")
            continue
        
        seen_descriptions.add(normalized_desc)
        
        item = {
            'description': description,
            'original_description': description,  # Add for compatibility with matching function
            'enhanced_description': enhanced_description,
            'quantity': quantity,
            'row_number': row_num,
            'sheet_name': sheet_name,
            'head_title': section_stack[-1] if section_stack else None,
            'section_context': ' > '.join(section_stack) if section_stack else 'General'
        }
        
        items.append(item)
        unique_items_added += 1
        logger.debug(f"Row {row_num}: Added unique item '{description[:30]}...' (qty: {quantity})")
    
    logger.info(f"Sheet '{sheet_name}' extraction summary:")
    logger.info(f"  Total rows scanned: {total_rows_scanned}")
    logger.info(f"  Data rows with quantities found: {data_rows_found}")
    logger.info(f"  Unique items extracted: {unique_items_added}")
    logger.info(f"  BOQ sections detected: {len(section_stack)}")
    logger.info(f"  Deduplication saved: {total_rows_scanned - unique_items_added} duplicates avoided")
    
    return items

def detect_section_header(description: str, quantity: Optional[float]) -> bool:
    """Detect if a row is a section header rather than an item - be more conservative"""
    desc_upper = description.upper().strip()
    
    # Only detect as section if very clear indicators
    if quantity is None or quantity <= 0:
        # Strong section indicators only
        strong_section_keywords = [
            'BILL NR', 'SUB-BILL', 'PART A', 'PART B', 'SECTION', 'CHAPTER', 'DIVISION'
        ]
        
        # Must be ALL CAPS and contain strong section keywords
        if desc_upper == description and len(description) > 10:
            if any(keyword in desc_upper for keyword in strong_section_keywords):
                return True
            
        # Very short headers that are clearly sections
        if len(description) <= 20 and any(word in desc_upper for word in ['SECTION', 'PART', 'CHAPTER', 'DIVISION']):
            return True
    
    return False

def update_section_stack(section_stack: List[str], section_title: str):
    """Update the hierarchical section stack"""
    # Clean the title
    clean_title = section_title.strip()
    
    # Determine hierarchy level based on formatting
    if clean_title.isupper() and len(clean_title) > 10:
        # Major section - clear stack and start fresh
        section_stack.clear()
        section_stack.append(clean_title)
    elif clean_title.isupper():
        # Subsection - keep one level
        if section_stack:
            section_stack = section_stack[:1]
        section_stack.append(clean_title)
    else:
        # Minor section or detail
        section_stack.append(clean_title)
        # Keep maximum 3 levels
        if len(section_stack) > 3:
            section_stack = section_stack[-3:]

def build_enhanced_description(description: str, section_stack: List[str]) -> str:
    """Build enhanced description with contextual information"""
    if not section_stack:
        return description
    
    # Create context string
    context = ' > '.join(section_stack)
    
    # Combine context with description
    enhanced = f"{context} > {description}"
    
    return enhanced

def should_skip_item(description: str, quantity: Optional[float] = None) -> bool:
    """Determine if an item should be skipped based on description patterns"""
    desc_lower = description.lower()
    
    # Skip obvious non-items
    skip_patterns = [
        'total', 'subtotal', 'grand total', 'sum', 'page', 'continued',
        'note:', 'remark:', 'see', 'refer to', 'as per', 'including',
        'excluding', 'carried forward', 'brought forward', 'page total'
    ]
    
    for pattern in skip_patterns:
        if pattern in desc_lower:
            return True
    
    # Skip if description is too short or generic - BUT be more lenient if there's a valid quantity
    min_length = 1 if (quantity is not None and quantity > 0) else 5
    if len(description.strip()) < min_length:
        return True
    
    return False

def detect_boq_section_header(description: str, quantity: Optional[float], row_num: int, ws) -> bool:
    """Enhanced BOQ section detection for different formats"""
    desc_upper = description.upper().strip()
    
    # Only consider as section if no quantity
    if quantity is not None and quantity > 0:
        return False
    
    # BOQ section patterns
    boq_section_patterns = [
        r'^BILL\s+N[RO]?\.?\s*\d+', # BILL NR 2005, BILL NO. 1, etc.
        r'^SUB[\-\s]?BILL', # SUB-BILL, SUB BILL
        r'^PART\s+[A-Z0-9]', # PART A, PART 1
        r'^SECTION\s+[A-Z0-9]', # SECTION A, SECTION 1
        r'^CHAPTER\s+[A-Z0-9]', # CHAPTER 1
        r'^[A-Z]\d{2,}', # A393, D20, etc.
        r'^\d+\.\d+', # 2005.02, 1.1, etc.
    ]
    
    # Check for BOQ patterns
    import re
    for pattern in boq_section_patterns:
        if re.match(pattern, desc_upper):
            return True
    
    # Common BOQ section headers
    boq_sections = [
        'GROUNDWORK', 'SUBSTRUCTURES', 'SUPERSTRUCTURE', 'ROOFING', 'EXTERNAL WALLS',
        'INTERNAL WALLS', 'FLOORS', 'STAIRS', 'ROOF', 'EXTERNAL DOORS', 'WINDOWS',
        'INTERNAL DOORS', 'WALL FINISHES', 'FLOOR FINISHES', 'CEILING FINISHES',
        'FITTINGS', 'DISPOSAL SYSTEMS', 'WATER INSTALLATIONS', 'HEAT SOURCE',
        'SPACE HEATING', 'VENTILATION', 'ELECTRICAL INSTALLATIONS', 'LIFT INSTALLATIONS',
        'PROTECTIVE INSTALLATIONS', 'COMMUNICATION INSTALLATIONS', 'SPECIAL INSTALLATIONS',
        'BUILDERS WORK', 'DRAINAGE', 'EXTERNAL WORKS', 'DEMOLITION', 'ALTERATIONS'
    ]
    
    # Check if it's a known BOQ section
    if any(section in desc_upper for section in boq_sections):
        return True
    
    # Check if it's all caps and looks like a header (but not just a short item description)
    if desc_upper == description and len(description) > 15:
        return True
    
    return False

def should_skip_boq_item(description: str, quantity: Optional[float], seen_descriptions: set) -> bool:
    """Enhanced filtering for BOQ items"""
    desc_lower = description.lower().strip()
    
    # Skip obvious non-items
    skip_patterns = [
        'total', 'subtotal', 'grand total', 'sum', 'page', 'continued',
        'note:', 'remark:', 'see', 'refer to', 'as per', 'including',
        'excluding', 'carried forward', 'brought forward', 'page total',
        'description', 'item', 'rate', 'amount', 'unit', 'qty', 'quantity'
    ]
    
    for pattern in skip_patterns:
        if pattern in desc_lower:
            return True
    
    # Skip very short or generic descriptions
    if len(description.strip()) < 2:
        return True
    
    # Skip single words that are likely column headers or non-descriptive
    words = description.strip().split()
    if len(words) == 1 and len(words[0]) < 10:
        generic_words = ['disposal', 'excavation', 'concrete', 'steel', 'timber', 'walls', 'doors', 'windows']
        if words[0].lower() in generic_words:
            return True
    
    return False

def normalize_description_for_dedup(description: str) -> str:
    """Normalize description for deduplication"""
    # Convert to lowercase and remove extra whitespace
    normalized = re.sub(r'\s+', ' ', description.lower().strip())
    
    # Remove common BOQ formatting characters
    normalized = re.sub(r'[^\w\s]', '', normalized)
    
    # Remove very common words that don't affect meaning
    stop_words = {'the', 'and', 'or', 'of', 'in', 'to', 'for', 'with', 'by', 'at', 'on', 'as', 'per'}
    words = [word for word in normalized.split() if word not in stop_words]
    
    return ' '.join(words)

def main():
    parser = argparse.ArgumentParser(description="Enhanced Cohere Excel Price Matching")
    parser.add_argument('--inquiry', required=True, help='Path to inquiry Excel file')
    parser.add_argument('--pricelist', required=True, help='Path to pricelist Excel file')
    parser.add_argument('--output', required=True, help='Path for output Excel file')
    parser.add_argument('--api-key', required=True, help='Cohere API key')
    parser.add_argument('--similarity-threshold', type=float, default=SIMILARITY_THRESHOLD, 
                       help='Minimum similarity threshold for matches')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose logging')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    try:
        logger.info("=== Enhanced Cohere Excel Price Matching Started ===")
        logger.info(f"Inquiry file: {args.inquiry}")
        logger.info(f"Pricelist file: {args.pricelist}")
        logger.info(f"Output file: {args.output}")
        logger.info(f"Similarity threshold: {args.similarity_threshold}")
        
        # Initialize Cohere client
        try:
            client = cohere.Client(args.api_key)
            logger.info("Cohere client initialized successfully")
        except Exception as e:
            raise Exception(f"Failed to initialize Cohere client: {str(e)}")
        
        # Load pricelist into DataFrame format for new processing
        price_descriptions, price_rates, price_units, price_ids = load_pricelist_enhanced(args.pricelist)
        
        # Create pricelist DataFrame
        pricelist_df = pd.DataFrame({
            'id': price_ids,
            'description': price_descriptions,
            'rate': price_rates,
            'unit': price_units
        })
        
        logger.info(f"Loaded pricelist with {len(pricelist_df)} items")
        
        # Use new multi-sheet processing
        job_id = str(uuid.uuid4())
        output_path = process_all_sheets(args.inquiry, pricelist_df, job_id, client)
        
        if output_path:
            # Copy output to specified path
            import shutil
            shutil.copy2(output_path, args.output)
            logger.info(f"Processing completed successfully! Output saved to: {args.output}")
        else:
            logger.error("Processing failed - no output generated")
            return 1
        
    except Exception as e:
        logger.error(f"Error in main: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return 1
    
    return 0

def process_item_matching(items_df: pd.DataFrame, pricelist_df: pd.DataFrame, job_id: str, client: cohere.Client) -> Optional[str]:
    """Process item matching and generate output Excel file"""
    try:
        logger.info("Cohere client ready for matching")
        
        # Prepare data for matching
        price_descriptions = pricelist_df['description'].tolist()
        price_rates = pricelist_df['rate'].tolist()
        price_units = pricelist_df['unit'].tolist()
        price_ids = pricelist_df['id'].tolist()
        
        # Preprocess descriptions
        processed_price_descs = [enhanced_preprocess(desc, SYNONYM_MAP, STOP_WORDS) 
                               for desc in price_descriptions]
        
        # Generate embeddings
        logger.info("Generating price embeddings...")
        price_embeddings = embed_texts_with_retry(client, processed_price_descs, "search_document")
        price_embeddings_norm = price_embeddings / np.linalg.norm(price_embeddings, axis=1, keepdims=True)
        
        # Process inquiry items
        logger.info(f"Processing {len(items_df)} inquiry items...")
        
        matches = []
        for idx, row in items_df.iterrows():
            # Use enhanced description for better matching
            inquiry_desc = row.get('enhanced_description', row.get('original_description', row.get('description', '')))
            # Ensure we have a valid string for preprocessing
            if inquiry_desc is None:
                inquiry_desc = str(row.get('original_description', row.get('description', '')))
            processed_inquiry = enhanced_preprocess(str(inquiry_desc), SYNONYM_MAP, STOP_WORDS)
            
            # Generate inquiry embedding
            inquiry_embedding = embed_texts_with_retry(client, [processed_inquiry], "search_query")
            inquiry_embedding_norm = inquiry_embedding / np.linalg.norm(inquiry_embedding)
            
            # Calculate similarities
            similarities = np.dot(price_embeddings_norm, inquiry_embedding_norm.T).flatten()
            
            # Find best match
            best_idx = np.argmax(similarities)
            best_similarity = similarities[best_idx]
            
            if best_similarity >= SIMILARITY_THRESHOLD:
                match = {
                    'id': str(uuid.uuid4()),
                    'original_description': row.get('original_description', row.get('description', '')),
                    'matched_description': price_descriptions[best_idx],
                    'matched_rate': price_rates[best_idx],
                    'similarity_score': float(best_similarity),
                    'row_number': int(row['row_number']),
                    'sheet_name': row['sheet_name'],
                    'quantity': float(row['quantity']),
                    'unit': price_units[best_idx],
                    'total_amount': float(row['quantity']) * price_rates[best_idx],
                    'matched_price_item_id': price_ids[best_idx],
                    'section_context': row.get('section_context', 'General')
                }
                matches.append(match)
                original_desc = row.get('original_description', row.get('description', ''))
                logger.debug(f"Matched: {str(original_desc)[:50]}... -> {price_descriptions[best_idx][:50]}... (sim: {best_similarity:.3f})")
            else:
                original_desc = row.get('original_description', row.get('description', ''))
                logger.debug(f"No match found for: {str(original_desc)[:50]}... (best sim: {best_similarity:.3f})")
        
        # Generate output Excel file compatible with existing JavaScript parser
        output_path = os.path.join('output', f'processed-{job_id}-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        os.makedirs('output', exist_ok=True)
        
        if matches:
            # Create Excel workbook with openpyxl directly for better control
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Define headers to match what JavaScript expects
            headers = [
                'original_description',
                'matched_description', 
                'matched_rate',
                'similarity_score',
                'quantity',
                'unit',
                'total_amount',
                'matched_price_item_id',
                'row_number',
                'sheet_name'
            ]
            
            # Add headers with formatting
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            # Add data rows
            for row_idx, match in enumerate(matches, 2):
                ws.cell(row=row_idx, column=1, value=match['original_description'])
                ws.cell(row=row_idx, column=2, value=match['matched_description'])
                ws.cell(row=row_idx, column=3, value=match['matched_rate'])
                ws.cell(row=row_idx, column=4, value=match['similarity_score'])
                ws.cell(row=row_idx, column=5, value=match['quantity'])
                ws.cell(row=row_idx, column=6, value=match['unit'])
                ws.cell(row=row_idx, column=7, value=match['total_amount'])
                ws.cell(row=row_idx, column=8, value=match['matched_price_item_id'])
                ws.cell(row=row_idx, column=9, value=match['row_number'])
                ws.cell(row=row_idx, column=10, value=match['sheet_name'])
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Generated output file with {len(matches)} matches: {output_path}")
            return output_path
        else:
            logger.warning("No matches found!")
            return None
            
    except Exception as e:
        logger.error(f"Error in process_item_matching: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

if __name__ == '__main__':
    main()
