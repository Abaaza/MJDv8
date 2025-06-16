"""pricing_engine.py
Reads a master price Excel sheet and returns a lookup dictionary.
This module relies on openpyxl; install via `pip install openpyxl`.
"""

from pathlib import Path
from openpyxl import load_workbook
import csv
import os
from datetime import datetime


def load_rates(path: str) -> dict:
    """Return a mapping of code -> {rate, cost} supporting CSV or Excel."""
    ext = Path(path).suffix.lower()
    rates: dict = {}

    if ext == ".csv":
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                code = row.get("code") or row.get("Code")
                if not code:
                    continue
                rate = row.get("rate") or row.get("Rate") or row.get("price") or row.get("Price")
                cost = row.get("cost") or row.get("Cost")
                rates[str(code).strip()] = {
                    "rate": float(rate or 0),
                    "cost": float(cost or 0),
                }
        return rates

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, rate, cost = row[0], row[1], None if len(row) < 3 else row[2]
        if code:
            rates[str(code).strip()] = {
                "rate": float(rate or 0),
                "cost": float(cost or 0),
            }
    return rates


def apply_rates(boq: list, rates: dict) -> dict:
    """Populate unit rates, costs and profit margin for BoQ items.

    Returns a dictionary with the priced items, overall total and any rate
    overrides detected during processing.
    """
    priced = []
    overrides = []

    for item in boq:
        data = rates.get(item.get("code"), {})
        rate = data.get("rate")
        cost = data.get("cost")
        unit_rate = item.get("unit_rate") if item.get("unit_rate") is not None else rate

        if (
            item.get("unit_rate") is not None
            and rate is not None
            and float(item["unit_rate"]) != float(rate)
        ):
            overrides.append(
                {
                    "timestamp": datetime.utcnow().isoformat(),
                    "code": item.get("code"),
                    "default_rate": rate,
                    "override_rate": item.get("unit_rate"),
                }
            )

        total = None
        profit = None
        margin = None
        if unit_rate is not None and item.get("qty") is not None:
            qty = float(item["qty"])
            total = float(unit_rate) * qty
            if cost is not None:
                profit = total - float(cost) * qty
                if total:
                    margin = (profit / total) * 100

        priced.append(
            {
                **item,
                "unit_rate": unit_rate,
                "total": total,
                "cost_rate": cost,
                "profit": profit,
                "margin": margin,
            }
        )

    job_total = sum(i["total"] for i in priced if i.get("total") is not None)

    return {"items": priced, "total": job_total, "overrides": overrides}

if __name__ == "__main__":
    import json
    import sys

    rates = load_rates(sys.argv[1])
    if len(sys.argv) > 2:
        with open(sys.argv[2]) as f:
            items = json.load(f)
    else:
        items = json.load(sys.stdin)
    result = apply_rates(items, rates)
    json.dump(result, sys.stdout, indent=2)